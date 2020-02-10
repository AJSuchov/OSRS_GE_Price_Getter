from openpyxl import load_workbook #Import workbook
from openpyxl import Workbook #Create and import workbook
import time
import decimal, random
from copy import copy, deepcopy
from operator import itemgetter
import bs4 as bs
import urllib.request, urllib.error

###########################################################################################
############# Retrieve excel file to read

# set file path to retrieve file
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Absence_Roster.xlsx"
#filepath="C:\\Users\\AJ Suchovsky\\Desktop\\Multiprocessing_Project\\Large_Roster_Test.xlsx"
filepath="C:\\Users\\suchovaj\\Documents\\Python Scripts\\OSRSList.xlsx"

# load demo.xlsx 
wb=load_workbook(filepath,data_only = True)

# select demo.xlsx
sheet=wb.active

# get max row count
max_row=sheet.max_row

# get max column count
max_column=sheet.max_column

################################################################################
def soupCollector(url, iterator, halt):
    try:
        time.sleep(halt)
        sauce = urllib.request.urlopen(url).read()
        soup = bs.BeautifulSoup(sauce,'lxml')
        title = soup.title.string
        #print(soup.find_all('h3'))
        if soup.title.string == "RuneScape Oldschool - Grand Exchange - Prices, Trade, Market Movers":
               soupCollector(url, iterator, halt) 
        else:
            for h3 in soup.find_all('h3'):
                print(title)
                print(h3.text + iterator)
    except urllib.error.HTTPError as e:
        print(e.__dict__)
    except urllib.error.URLError as e:
        print(e.__dict__)

def retriever(start,end):
    alist = []
    
    for i in range(start,end):
        #time.sleep(0.5)
    # iterate over all columns
        blist = []
        for j in range(1,max_column+1):
          # get particular cell value    
            cell_obj=sheet.cell(row=i,column=j)
          # print cell value     
          #list.append(cell_obj.value)
            blist.append(cell_obj.value)
        alist.append(blist)
     # print new line
    return alist

if __name__ == '__main__':
    itemInfo = retriever(1,max_row+1)

    i = 0
    j = 0

    for item in itemInfo:
        i=i+1
##        if j == 21:
##            time.sleep(3)
##            j = 0
##            print(str(j) + " This is reset to 0 and wait 3 seconds")
        range_num = float(decimal.Decimal(random.randrange(5,10)))
        url = item[2]
        soupCollector(url,str(i), range_num)

